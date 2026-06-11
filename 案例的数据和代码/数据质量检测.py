import pandas as pd
import numpy as np
import os  # 新增：用于自动获取桌面路径
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ===================== 1. 核心配置项（已适配桌面路径）=====================
# 自动获取你电脑的桌面路径（Windows/Mac通用）
desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

# 输入文件：读取桌面上的「diabetic_data - 副本.xlsx」
INPUT_FILE = os.path.join(desktop_path, "diabetic_data - 副本.xlsx")
# 输出文件：生成的报告直接保存到桌面
OUTPUT_FILE = os.path.join(desktop_path, "糖尿病并发症数据_质量检测报告.xlsx")

# 目标工作表名称
SHEET_NAME = 'diabetic_data'

# 字段合法值规则（完全匹配你指定的编码规则）
VALID_VALUE_RULES = {
    '性别': {
        'valid_values': [0, 1,2],
        'description': '女→0，男→1，Unknow→2，仅允许0/1/2'
    },
    '种族': {
        'valid_values': [0, 1, 2, 3, 4, 5],
        'description': '白种人→0，非裔美国人→1，亚裔→2，西班牙裔→3，Other→4，空/NaN→5，仅允许0-5'
    },
    '年龄': {
        'valid_values': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
        'description': '[0-10)→1 至 [90-100)→10，仅允许1-10的整数'
    },
    '再次入院情况': {
        'valid_values': [0, 1, 2],
        'description': 'NO→0，>30→1，<30→2，仅允许0/1/2'
    },
    '最高血清血糖': {
        'valid_values': [0, 1, 2, 3],
        'description': 'None→0，>200→1，>300→2，Norm→3，仅允许0/1/2/3'
    },
    '糖化血红蛋白结果': {
        'valid_values': [0, 1, 2, 3],
        'description': 'None→0，>7→1，>8→2，Norm→3，仅允许0/1/2/3'
    },
    'has_complication': {
        'valid_values': [0, 1],
        'description': '是否有并发症，仅允许0/1二值'
    }
}

# 数值型字段（用于异常值检测）
NUMERIC_COLS = [
    '住院天数', '实验室检查次数', '医疗操作次数', '使用药物数量',
    '门诊就诊次数', '急诊就诊次数', '住院次数', '诊断总数'
]

# 异常值检测阈值（3σ原则）
SIGMA_THRESHOLD = 3

# ===================== 2. 数据加载 =====================
print("="*50)
print("1. 数据加载中...")
# 读取数据
df = pd.read_excel(INPUT_FILE, sheet_name=SHEET_NAME)
total_rows, total_cols = df.shape
print(f"数据加载完成，总数据量：{total_rows}行 × {total_cols}列")
print("="*50)

# ===================== 3. 核心质量检测模块 =====================
# 3.1 数据基本信息检测
print("\n2. 数据基本信息检测")
basic_info = pd.DataFrame({
    '指标': ['总行数', '总列数', '字段列表', '数据类型分布'],
    '值': [
        total_rows,
        total_cols,
        '、'.join(df.columns.tolist()),
        str(df.dtypes.value_counts().to_dict())
    ]
})
print(basic_info)
print("="*50)

# 3.2 缺失值检测
print("\n3. 缺失值检测")
missing_stats = pd.DataFrame({
    '字段名': df.columns,
    '缺失值数量': df.isnull().sum(),
    '缺失值占比(%)': (df.isnull().sum() / total_rows * 100).round(4)
}).reset_index(drop=True)
# 按缺失值数量降序排序
missing_stats = missing_stats.sort_values('缺失值数量', ascending=False).reset_index(drop=True)
print(missing_stats)
print("="*50)

# 3.3 重复值检测
print("\n4. 重复值检测")
duplicate_count = df.duplicated().sum()
duplicate_stats = pd.DataFrame({
    '指标': ['总数据行数', '重复数据行数', '重复数据占比(%)', '去重后数据行数'],
    '值': [
        total_rows,
        duplicate_count,
        round(duplicate_count / total_rows * 100, 4),
        total_rows - duplicate_count
    ]
})
print(duplicate_stats)
print("="*50)

# 3.4 枚举值合规性检测（核心校验）
print("\n5. 枚举值合规性检测")
compliance_results = []
invalid_detail = {}

for col_name, rule in VALID_VALUE_RULES.items():
    if col_name not in df.columns:
        compliance_results.append({
            '字段名': col_name,
            '规则说明': rule['description'],
            '合法值范围': str(rule['valid_values']),
            '总数据行数': total_rows,
            '合规行数': 0,
            '不合规行数': 0,
            '合规率(%)': 0,
            '状态': '字段不存在'
        })
        continue
    
    # 计算合规/不合规数量
    valid_mask = df[col_name].isin(rule['valid_values'])
    valid_count = valid_mask.sum()
    invalid_count = total_rows - valid_count
    valid_rate = round(valid_count / total_rows * 100, 4)
    
    # 记录不合规值详情
    if invalid_count > 0:
        invalid_values = df[~valid_mask][col_name].value_counts().to_dict()
        invalid_detail[col_name] = invalid_values
        status = '不合规'
    else:
        status = '合规'
    
    compliance_results.append({
        '字段名': col_name,
        '规则说明': rule['description'],
        '合法值范围': str(rule['valid_values']),
        '总数据行数': total_rows,
        '合规行数': valid_count,
        '不合规行数': invalid_count,
        '合规率(%)': valid_rate,
        '状态': status
    })

# 转换为DataFrame
compliance_df = pd.DataFrame(compliance_results)
print(compliance_df)

# 打印不合规值详情
if invalid_detail:
    print("\n不合规值详情：")
    for col_name, values in invalid_detail.items():
        print(f"\n{col_name} 不合规值分布：{values}")
print("="*50)

# 3.5 数值型字段异常值检测（3σ原则）
print("\n6. 数值型字段异常值检测")
outlier_results = []

for col_name in NUMERIC_COLS:
    if col_name not in df.columns:
        outlier_results.append({
            '字段名': col_name,
            '总数据行数': total_rows,
            '均值': np.nan,
            '标准差': np.nan,
            '3σ下限': np.nan,
            '3σ上限': np.nan,
            '异常值数量': 0,
            '异常值占比(%)': 0,
            '状态': '字段不存在'
        })
        continue
    
    # 计算3σ范围
    col_data = df[col_name].dropna()
    mean_val = col_data.mean()
    std_val = col_data.std()
    lower_bound = mean_val - SIGMA_THRESHOLD * std_val
    upper_bound = mean_val + SIGMA_THRESHOLD * std_val
    
    # 计算异常值数量
    outlier_mask = (df[col_name] < lower_bound) | (df[col_name] > upper_bound)
    outlier_count = outlier_mask.sum()
    outlier_rate = round(outlier_count / total_rows * 100, 4)
    
    status = '有异常值' if outlier_count > 0 else '无异常值'
    
    outlier_results.append({
        '字段名': col_name,
        '总数据行数': total_rows,
        '均值': round(mean_val, 4),
        '标准差': round(std_val, 4),
        '3σ下限': round(lower_bound, 4),
        '3σ上限': round(upper_bound, 4),
        '异常值数量': outlier_count,
        '异常值占比(%)': outlier_rate,
        '状态': status
    })

# 转换为DataFrame
outlier_df = pd.DataFrame(outlier_results)
print(outlier_df)
print("="*50)

# ===================== 4. 生成Excel质量检测报告 =====================
print("\n7. 生成Excel质量检测报告...")

# 创建工作簿
wb = Workbook()

# 4.1 报告总览Sheet
ws_overview = wb.active
ws_overview.title = '质量检测总览'

# 标题样式
title_font = Font(bold=True, size=14, color='FFFFFF')
title_fill = PatternFill('solid', start_color='4472C4')
center_align = Alignment(horizontal='center', vertical='center')

# 写入标题
ws_overview.merge_cells('A1:H1')
title_cell = ws_overview['A1']
title_cell.value = '糖尿病并发症数据质量检测总报告'
title_cell.font = title_font
title_cell.fill = title_fill
title_cell.alignment = center_align

# 写入基本信息
basic_data = [
    ['检测时间', pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')],
    ['数据文件', INPUT_FILE],
    ['总数据行数', total_rows],
    ['总字段数', total_cols],
    ['缺失字段总数', len(missing_stats[missing_stats['缺失值数量'] > 0])],
    ['重复数据行数', duplicate_count],
    ['不合规字段总数', len(compliance_df[compliance_df['状态'] == '不合规'])],
    ['有异常值字段总数', len(outlier_df[outlier_df['状态'] == '有异常值'])]
]

# 写入基本信息到Sheet
for i, (key, value) in enumerate(basic_data, start=3):
    ws_overview.cell(row=i, column=1, value=key).font = Font(bold=True)
    ws_overview.cell(row=i, column=2, value=value)

# 4.2 缺失值检测Sheet
ws_missing = wb.create_sheet('缺失值检测')
# 写入表头
missing_headers = missing_stats.columns.tolist()
for col_idx, header in enumerate(missing_headers, start=1):
    cell = ws_missing.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True)
    cell.alignment = center_align
# 写入数据
for row_idx, row_data in enumerate(missing_stats.values, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws_missing.cell(row=row_idx, column=col_idx, value=value)

# 4.3 枚举值合规性检测Sheet
ws_compliance = wb.create_sheet('枚举值合规性检测')
# 写入表头
compliance_headers = compliance_df.columns.tolist()
for col_idx, header in enumerate(compliance_headers, start=1):
    cell = ws_compliance.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True)
    cell.alignment = center_align
# 写入数据
for row_idx, row_data in enumerate(compliance_df.values, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_compliance.cell(row=row_idx, column=col_idx, value=value)
        # 不合规行标红
        if row_data[-1] == '不合规':
            cell.font = Font(color='FF0000')

# 4.4 异常值检测Sheet
ws_outlier = wb.create_sheet('数值异常值检测')
# 写入表头
outlier_headers = outlier_df.columns.tolist()
for col_idx, header in enumerate(outlier_headers, start=1):
    cell = ws_outlier.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True)
    cell.alignment = center_align
# 写入数据
for row_idx, row_data in enumerate(outlier_df.values, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws_outlier.cell(row=row_idx, column=col_idx, value=value)
        # 有异常值行标红
        if row_data[-1] == '有异常值':
            cell.font = Font(color='FF0000')

# 4.5 自动调整列宽
for ws in wb.worksheets:
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
        # 设置列宽，最大不超过50
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width

# 保存文件
wb.save(OUTPUT_FILE)
print(f"质量检测报告已生成，保存路径：{OUTPUT_FILE}")
print("="*50)
print("所有质量检测任务执行完成！")
